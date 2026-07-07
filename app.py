import streamlit as st
import pandas as pd
import random
import copy
import threading
import queue
import urllib.parse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

# =========================================================================
# 系統公告留言板（寫死在 Code 裡，留空 "" 則不啟用）
# =========================================================================
# 📝 如果要發公告，請在下方引號內打字；如果要關閉，請保持 ANNOUNCEMENT = ""
ANNOUNCEMENT = """
📣 **系統維護與排班公告**
1. 本週已調整「偏好 6A 員工」的特殊班權重分配邏輯。
2. 若資深群組同步失敗，請檢查 Sheet 頁籤名稱是否遭誤動。
3. 有任何排班邏輯建議，請隨時通知岱川。
"""

# 定義彈窗模組
@st.dialog("📌 系統重要備註與公告")
def show_announcement_dialog(text):
    st.markdown(text)
    if st.button("我知道了，關閉提示", type="primary"):
        st.rerun()

# 初始化公告檢查狀態
if "announcement_shown" not in st.session_state:
    st.session_state["announcement_shown"] = False

# 邏輯：Code 裡有打字，且這次進來還沒顯示過，就立刻跳出
if ANNOUNCEMENT.strip() and not st.session_state["announcement_shown"]:
    st.session_state["announcement_shown"] = True
    show_announcement_dialog(ANNOUNCEMENT)

# =========================================================================


# 頁面基本設定
st.set_page_config(page_title="排班系統", page_icon="📅", layout="wide")

st.title("📅 排班系統 ")
st.markdown("側邊欄可調整參數；中間主畫面選擇群組並同步後，即可執行背景排班。")

# =========================================================================
# 側邊欄：參數設定區
# =========================================================================
st.sidebar.header("⚙️ 基礎班別每月需求人數")
need_A = st.sidebar.number_input("A 班每月需求人數", min_value=0, value=18)
need_E = st.sidebar.number_input("E 班每月需求人數", min_value=0, value=17)
need_N = st.sidebar.number_input("N 班每月需求人數", min_value=0, value=12)
shift_needs = {"A": need_A, "E": need_E, "N": need_N}

st.sidebar.markdown("---")
st.sidebar.header("🎲 特殊班別 (W/X/Y/Z) 每月配額")
need_W = st.sidebar.slider("W 班數量 (A 類特殊)", 0, 10, 2)
need_X = st.sidebar.slider("X 班數量 (A 類特殊)", 0, 10, 2)
need_Y = st.sidebar.slider("Y 班數量 (E 類特殊)", 0, 10, 2)
need_Z = st.sidebar.slider("Z 班數量 (N 類特殊)", 0, 10, 2)
special_needs = {"W": need_W, "X": need_X, "Y": need_Y, "Z": need_Z}

st.sidebar.markdown("---")
st.sidebar.header("🛡️ 個人上限與保底設定")
max_special_normal = st.sidebar.slider("一般員工特殊班上限（每人）", 1, 5, 2)
max_special_6A = st.sidebar.slider("偏好 6A 員工特殊班上限（每人）", 0, 5, 1)
max_attempts = st.sidebar.number_input("五門檻保底最大嘗試次數", min_value=100, value=1000, step=100)

# 【已修正】完美符合直覺的「上、下」半年排序
st.sidebar.markdown("---")
st.sidebar.header("📅 排班季度選擇")
season_option = st.sidebar.selectbox(
    "請選擇排班月份區間：",
    ["上半年 (Jan - Jun)", "下半年 (Jul - Dec)"]
)

if "上半年" in season_option:
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]
else:
    months = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

# =========================================================================
# 側邊欄：開發者調試區（暗碼保護固定 Seed）
# =========================================================================
st.sidebar.markdown("---")
with st.sidebar.expander("🛠️ 開發者內部調試區", expanded=False):
    dev_password = st.text_input("輸入管理員密碼：", type="password")
    
    if dev_password == "117493":  # 這裡設定你的專屬密碼
        st.success("🔓 已啟用開發者模式：排班隨機碼已固定 (Seed=42)")
        random.seed(42)  # 鎖死隨機碼
    else:
        # 🔑 關鍵解鎖點：只要密碼不對或留白，立刻恢復真正隨機！
        random.seed(None) 
        if dev_password != "":
            st.error("❌ 密碼錯誤，維持正常隨機排班")

# =========================================================================
# 主畫面：暗碼安全資料串接
# =========================================================================
st.header("🔗 第一步：同步 Google Sheet 偏好資料")

col_url, col_opt = st.columns([2, 1])

with col_url:
    try:
        default_url = st.secrets["private_gsheets_url"]
    except:
        default_url = ""
    sheet_url = st.text_input("Google Sheet 網址 ：", value="")

with col_opt:
    group_option = st.selectbox("請選擇目前要處理的年資群組：", ["資深群組 (分頁1)", "中生代群組 (分頁2)", "新進群組 (分頁3)"])

sheet_mapping = {"資深群組 (分頁1)": "工作表1", "中生代群組 (分頁2)": "工作表2", "新進群組 (分頁3)": "工作表3"}
target_sheet_name = sheet_mapping[group_option]

employees = {}

if st.button("🔄 同步該群組最新資料", type="secondary"):
    active_url = sheet_url if sheet_url else default_url
    
    if not active_url:
        st.error("❌ 請輸入 Google Sheet 網址或在後台 Secrets 中設定 private_gsheets_url 網址。")
    else:
        try:
            # =========================================================================
            # 【一勞永逸終極通用版】自動適應新表、防空行、三層防呆機制
            # =========================================================================
            if "docs.google.com/spreadsheets" in active_url:
                base_url = active_url.split("/edit")[0]
                encoded_sheet_name = urllib.parse.quote(target_sheet_name)
                final_csv_url = f"{base_url}/gviz/tq?tqx=out:csv&sheet={encoded_sheet_name}"
            else:
                final_csv_url = active_url
            
            try:
                # 讀取 CSV 資料
                raw_df = pd.read_csv(final_csv_url)
                
                # 【防禦空行】強制濾除 name 或 ID 為空值的「幽靈列」
                if 'name' in raw_df.columns and 'ID' in raw_df.columns:
                    raw_df = raw_df.dropna(subset=['name', 'ID'])
                else:
                    raise ValueError("欄位名稱不符")
                    
            except Exception as e:
                # 第二層防呆：當上方的精準分頁通道失敗時（例如別人改了分頁名字），自動降級抓第一個分頁
                fallback_url = f"{active_url.split('/edit')[0]}/export?format=csv&gid=0"
                raw_df = pd.read_csv(fallback_url)
                if 'name' in raw_df.columns and 'ID' in raw_df.columns:
                    raw_df = raw_df.dropna(subset=['name', 'ID'])
                st.warning(f"⚠️ 讀取分頁【{target_sheet_name}】失敗，系統已自動為您切換至第一個預設工作表。請檢查 Google Sheet 分頁名稱是否真的叫「工作表1/2/3」。")
            
            # 【關鍵修復】確保資料清理乾淨後，再建立唯一的 key
            raw_df['key'] = raw_df['name'].astype(str) + '_' + raw_df['ID'].astype(str).str.split('.').str[0]
            missing_months = [m for m in months if m not in raw_df.columns]
            
            if missing_months:
                st.error(f"❌ 同步失敗！目前選擇的是【{season_option}】，但你的 Google Sheet 內找不到欄位：{missing_months}。請檢查 Sheet 欄位名稱或切換季度。")
            else:
                invalid_pref_records = []
                pref_df_cleaned = raw_df.set_index('key')[months].copy()

                for emp in pref_df_cleaned.index:
                    for i, m in enumerate(months):
                        raw = str(pref_df_cleaned.at[emp, m]).strip().upper()
                        if raw not in {"A", "E", "N"}:
                            invalid_pref_records.append((emp, m, raw))
                            pref_df_cleaned.at[emp, m] = "A"

                temp_employees = pref_df_cleaned.T.to_dict('list')

                # 【校對修正】將演算順序修正，確保 A_pref_counts 讀取到正確建立後的 temp_employees
                A_pref_counts = {i: 0 for i in range(len(months))}
                for pref in temp_employees.values():
                    for i, v in enumerate(pref):
                        if v == "A": A_pref_counts[i] += 1

                for emp, pref in temp_employees.items():
                    if "A" not in pref:
                        min_count = min(A_pref_counts.values())
                        candidate_idxs = [i for i, cnt in A_pref_counts.items() if cnt == min_count]
                        chosen_idx = random.choice(candidate_idxs)
                        pref[chosen_idx] = "A"
                        temp_employees[emp] = pref
                        A_pref_counts[chosen_idx] += 1

                st.session_state['loaded_employees'] = temp_employees
                st.session_state['invalid_records'] = invalid_pref_records
                st.session_state['current_group'] = group_option
                st.success(f"✅ 成功同步【{group_option}】！共讀取 {len(temp_employees)} 人。")
                st.dataframe(raw_df.set_index('key')[months])
        except Exception as e:
            st.error(f"❌ 讀取失敗！請確認 Google Sheet 的共用權限已開啟為「知道連結的使用者皆可檢視」。錯誤: {e}")

if 'loaded_employees' in st.session_state:
    employees = st.session_state['loaded_employees']

# =========================================================================
# 核心排班與 Excel 導出邏輯
# =========================================================================
def assign_shifts(employees, shift_needs, months):
    final_df = pd.DataFrame(index=employees.keys(), columns=months)
    locked_A_employees = [emp for emp, pref in employees.items() if pref == ['A'] * len(months)]
    for emp in locked_A_employees:
        for m in months: final_df.at[emp, m] = "A"
    for i, m in enumerate(months):
        current = final_df[m].value_counts().to_dict()
        need_A = shift_needs["A"] - current.get("A", 0)
        need_E = shift_needs["E"] - current.get("E", 0)
        need_N = shift_needs["N"] - current.get("N", 0)
        unassigned = final_df[m][final_df[m].isnull()].index.tolist()
        def assign_by_pref(need, target):
            if need <= 0: return
            cand = [e for e in unassigned if employees[e][i] == target]
            if not cand: return
            pick = random.sample(cand, min(len(cand), need))
            for e in pick: final_df.at[e, m] = target; unassigned.remove(e)
        assign_by_pref(need_A, "A"); assign_by_pref(need_E, "E"); assign_by_pref(need_N, "N")
        remaining = final_df[m][final_df[m].isnull()].index.tolist()
        total_needed = (["A"] * max(0, shift_needs["A"] - final_df[m].value_counts().get("A", 0)) + ["E"] * max(0, shift_needs["E"] - final_df[m].value_counts().get("E", 0)) + ["N"] * max(0, shift_needs["N"] - final_df[m].value_counts().get("N", 0)))
        for e, s in zip(remaining, total_needed): final_df.at[e, m] = s
    for emp in final_df.index:
        if (final_df.loc[emp] == "A").sum() == 0:
            preferred_A_months = [months[i] for i in range(len(months)) if employees[emp][i] == "A"]
            candidate_months = preferred_A_months + months; random.shuffle(candidate_months)
            assigned = False
            for m in candidate_months:
                if final_df.at[emp, m] == "A": assigned = True; break
                A_emps = final_df[final_df[m] == "A"].index.tolist()
                swap_target = None
                for other in A_emps:
                    if employees[other] == ['A']*len(months): continue
                    if employees[other][months.index(m)] != "A" and (final_df.loc[other] == "A").sum() > 1: swap_target = other; break
                if not swap_target:
                    cand = [e for e in A_emps if employees[e] != ['A']*len(months) and (final_df.loc[e] == "A").sum() > 1]
                    if cand: swap_target = random.choice(cand)
                if swap_target:
                    ori = final_df.at[emp, m]; final_df.at[emp, m] = "A"; final_df.at[swap_target, m] = ori; assigned = True; break
    return final_df

def calculate_weighted_scores(schedule_df, employees, months, shift_needs):
    scores = {}
    pref_counts = []
    for i, m in enumerate(months):
        pref_counts.append({"A": sum(1 for e in employees if employees[e][i] == "A"), "E": sum(1 for e in employees if employees[e][i] == "E"), "N": sum(1 for e in employees if employees[e][i] == "N")})
    for emp in schedule_df.index:
        total = 0
        for i, m in enumerate(months):
            assigned = schedule_df.at[emp, m]
            _SPECIAL_SHIFT_MAP = {"W": "A", "X": "A", "Y": "E", "Z": "N"}
            assigned = _SPECIAL_SHIFT_MAP.get(assigned, assigned)
            preferred = employees[emp][i]
            if assigned == preferred:
                need = shift_needs[assigned]; pc = pref_counts[i][assigned]
                total += (1 if pc > need else 0.7)
        scores[emp] = round(total, 1)
    return scores

def perform_swap_weighted(schedule_df, ghost, lucky, months, employees, shift_needs, swap_history):
    for i, m in enumerate(months):
        key = tuple(sorted([ghost, lucky]) + [m])
        if key in swap_history: continue
        ghost_shift = schedule_df.at[ghost, m]; lucky_shift = schedule_df.at[lucky, m]; ghost_pref = employees[ghost][i]
        if ghost_shift == "A" and (schedule_df.loc[ghost] == "A").sum() == 1: continue
        if lucky_shift == ghost_pref and ghost_shift != ghost_pref:
            schedule_df.at[ghost, m], schedule_df.at[lucky, m] = lucky_shift, ghost_shift
            counts = schedule_df[m].value_counts().to_dict()
            if all(counts.get(s, 0) == shift_needs[s] for s in shift_needs): swap_history.add(key); return True
            schedule_df.at[ghost, m], schedule_df.at[lucky, m] = ghost_shift, lucky_shift
    return False

def ensure_at_least_one_A(schedule_df, employees, months, shift_needs):
    no_A = [e for e in schedule_df.index if (schedule_df.loc[e] == "A").sum() == 0]
    random.shuffle(no_A)
    for emp in no_A:
        assigned = False; pref_A_months = []
        for i, m in enumerate(months):
            if employees[emp][i] == "A" and sum(1 for e in employees if employees[e][i] == "A") <= shift_needs["A"]: pref_A_months.append(m)
        other = [m for m in months if m not in pref_A_months]; random.shuffle(other); candidates = pref_A_months + other
        for m in candidates:
            A_emps = schedule_df[schedule_df[m] == "A"].index.tolist(); random.shuffle(A_emps)
            for o in A_emps:
                if employees[o][months.index(m)] != "A" and (schedule_df.loc[o] == "A").sum() > 1:
                    ori = schedule_df.at[emp, m]; schedule_df.at[emp, m] = "A"; schedule_df.at[o, m] = ori; assigned = True; break
            if assigned: break
        if not assigned:
            for m in candidates:
                A_emps = schedule_df[schedule_df[m] == "A"].index.tolist(); random.shuffle(A_emps)
                for o in A_emps:
                    if employees[o] == ['A']*len(months):
                        ori = schedule_df.at[emp, m]; schedule_df.at[emp, m] = "A"; schedule_df.at[o, m] = ori; assigned = True; break
                if assigned: break
    return schedule_df

def apply_weighted_minimum_match(schedule_df, employees, months, shift_needs, max_attempts):
    precise = calculate_weighted_scores(schedule_df, employees, months, shift_needs); success = False
    for min_score in [4.5, 4.0, 3.5, 3.0, 2.5]:
        attempt, swap_history = 0, set()
        while attempt < max_attempts:
            attempt += 1
            ghosts = sorted([e for e, s in precise.items() if s < min_score], key=lambda x: precise[x])
            if not ghosts: success = True; break
            lucky = [e for e, s in precise.items() if s >= min_score and employees[e] != ['A']*len(months)]
            swap_ok = False
            for g in ghosts:
                for l in lucky:
                    if perform_swap_weighted(schedule_df, g, l, months, employees, shift_needs, swap_history):
                        precise = calculate_weighted_scores(schedule_df, employees, months, shift_needs); swap_ok = True; break
                if swap_ok: break
            if not swap_ok: break
        if success: break
    return ensure_at_least_one_A(schedule_df, employees, months, shift_needs)

def assign_special_shifts(schedule_df, months, special_needs, max_special_per_emp, max_special_6a, employees, shift_needs):
    final_df = schedule_df.copy(); special_shift_map = {"W": "A", "X": "A", "Y": "E", "Z": "N"}
    scores = calculate_weighted_scores(final_df, employees, months, shift_needs); special_count = {e: 0 for e in final_df.index}
    sixA_people = [e for e in final_df.index if employees[e] == ['A']*len(months)]; shuffled_months = months[:]
    random.shuffle(shuffled_months)
    for m in shuffled_months:
        for sp, base in special_shift_map.items():
            need = special_needs.get(sp, 0)
            if need <= 0: continue
            for slot_idx in range(need):
                candidates = final_df[final_df[m] == base].index.tolist()
                primary = [e for e in candidates if e not in sixA_people and special_count[e] < max_special_per_emp and final_df.loc[e].tolist().count(base) > 1]
                chosen = None
                if primary:
                    primary_sorted = sorted(primary, key=lambda e: (-scores.get(e, 0), special_count[e], random.random()))
                    chosen = primary_sorted[0]
                else:
                    backup = [e for e in candidates if e in sixA_people and special_count[e] < max_special_6a and final_df.loc[e].tolist().count(base) > 1]
                    if backup: chosen = random.choice(backup)
                if chosen: final_df.at[chosen, m] = sp; special_count[chosen] += 1
    return final_df

def run_scheduling_worker(employees, shift_needs, months, max_attempts, special_needs, max_special_normal, max_special_6A, res_queue):
    try:
        initial_df = assign_shifts(employees, shift_needs, months)
        match_df = apply_weighted_minimum_match(initial_df, employees, months, shift_needs, max_attempts)
        final_df = assign_special_shifts(match_df, months, special_needs, max_special_normal, max_special_6A, employees, shift_needs)
        res_queue.put(("SUCCESS", final_df))
    except Exception as e: res_queue.put(("ERROR", str(e)))

def generate_excel_bytes(schedule_df, employees, months, shift_needs, is_check_version=False):
    _SPECIAL_SHIFT_STYLE = {"W": ("A", Font(name="Arial", italic=True, bold=True)), "X": ("A", Font(name="Arial", underline="single", bold=True)), "Y": ("E", Font(name="Arial", underline="single", bold=True)), "Z": ("N", Font(name="Arial", underline="single", bold=True))}
    wb = Workbook(); ws = wb.active; ws.title = "排班結果"
    ws.cell(row=1, column=1, value="員工").font = Font(name="Arial", bold=True)
    for c, m in enumerate(months, start=2): ws.cell(row=1, column=c, value=m).font = Font(name="Arial", bold=True)
    if is_check_version:
        ws.cell(row=1, column=len(months)+2, value="命中數").font = Font(name="Arial", bold=True)
        ws.cell(row=1, column=len(months)+3, value="加權分數").font = Font(name="Arial", bold=True)
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"); scores = calculate_weighted_scores(schedule_df, employees, months, shift_needs)
    for r, emp in enumerate(schedule_df.index, start=2):
        ws.cell(row=r, column=1, value=emp).alignment = Alignment(horizontal="center", vertical="center")
        hit_count = 0
        for c, m in enumerate(months, start=2):
            val = str(schedule_df.at[emp, m]); cell = ws.cell(row=r, column=c)
            if val in _SPECIAL_SHIFT_STYLE:
                disp, style = _SPECIAL_SHIFT_STYLE[val]; cell.value = disp; cell.font = style
            else: cell.value = val; cell.font = Font(name="Arial")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            _SPECIAL_SHIFT_MAP = {"W": "A", "X": "A", "Y": "E", "Z": "N"}
            final_shift = _SPECIAL_SHIFT_MAP.get(val, val); preferred = str(employees[emp][c - 2])
            if final_shift == preferred: hit_count += 1
            elif is_check_version: cell.fill = yellow
        if is_check_version:
            ws.cell(row=r, column=len(months)+2, value=hit_count).alignment = Alignment(horizontal="center")
            ws.cell(row=r, column=len(months)+3, value=scores.get(emp, 0)).alignment = Alignment(horizontal="center")
    for c in range(1, ws.max_column + 1):
        col_letter = get_column_letter(c); max_len = max(len(str(ws.cell(row=r, column=c).value or '')) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[col_letter].width = max(25, max_len * 2 + 6) if col_letter == "A" else max(12, max_len * 2 + 3)
    output = io.BytesIO(); wb.save(output); return output.getvalue()

# =========================================================================
# 介面控制：執行排班與下載（真・異步安全線程版）
# =========================================================================
st.markdown("---")
st.header("🚀 第二步：執行智能排班")

# 初始化線程與隊列的狀態
if "thread_queue" not in st.session_state:
    st.session_state["thread_queue"] = None
if "is_running" not in st.session_state:
    st.session_state["is_running"] = False

if not employees:
    st.info("💡 請先在上方點擊「🔄 同步該群組最新資料」按鈕。")
else:
    st.write(f"📊 目前準備排班之群組：**{st.session_state.get('current_group', '未指定')}** ｜ 月份區間：**{season_option}**")
    
    # 狀態 1：如果目前沒有在排班，顯示「開始排班」按鈕
    if not st.session_state["is_running"]:
        if st.button("🔥 開始一鍵排班（真．背景異步）", type="primary"):
            res_queue = queue.Queue()
            task_thread = threading.Thread(
                target=run_scheduling_worker,
                args=(employees, shift_needs, months, max_attempts, special_needs, max_special_normal, max_special_6A, res_queue)
            )
            # 啟動背景工人
            task_thread.start()
            
            # 將 queue 記在 session_state，並標記開始執行，直接 rerun 釋放畫面
            st.session_state["thread_queue"] = res_queue
            st.session_state["is_running"] = True
            st.rerun()

    # 狀態 2：如果背景正在計算，顯示動態動畫，並檢查結果
    else:
        st.info("🧠 演算法正在後台隨機保底、分攤特殊班... 請稍候...")
        st.spinner("背景計算中...")
        
        # 嘗試從隊列拿東西，拿不到就代表還沒算完
        try:
            res_queue = st.session_state["thread_queue"]
            # block=False 代表不等待，有東西就拿，沒東西就噴 queue.Empty 異常
            status, result = res_queue.get(block=False)
            
            # 算完了！恢復狀態
            st.session_state["is_running"] = False
            st.session_state["thread_queue"] = None
            
            if status == "SUCCESS":
                st.session_state['final_result'] = result
                st.success("🎉 排班順利完成！結果已生成。")
            else:
                st.error(f"❌ 排班失敗: {result}")
            st.rerun()
            
        except queue.Empty:
            # 沒算完，利用 st.button 讓使用者可以手動刷新，或者放著等它下一輪觸發
            if st.button("🔄 檢查最新排班進度"):
                st.rerun()

if 'final_result' in st.session_state:
    st.markdown("### 📊 本次排班結果預覽")
    st.dataframe(st.session_state['final_result'], use_container_width=True)
    
    st.markdown("### 💾 下載報表")
    col1, col2 = st.columns(2)
    final_df = st.session_state['final_result']
    group_name = st.session_state.get('current_group', '群組')
    
    with col1:
        st.download_button(
            label="📥 下載【乾淨發布版】Excel",
            data=generate_excel_bytes(final_df, employees, months, shift_needs, is_check_version=False),
            file_name=f"抽班結果_{group_name}_{season_option}_發布版.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col2:
        st.download_button(
            label="📥 下載【後台檢查版】(含黃底未命中+加權分)",
            data=generate_excel_bytes(final_df, employees, months, shift_needs, is_check_version=True),
            file_name=f"抽班結果_{group_name}_{season_option}_檢查版.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
